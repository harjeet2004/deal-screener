"""
Generate 5 fake portfolio-company Excel files with QUARTERLY data (FY23 Q1–FY25 Q4).
12 quarters per company (3 financial years). Formats deliberately inconsistent across
companies — different quarter-label styles, junk rows, column orderings.

Numbers are in Indian Rupees at realistic seed-to-Series-B scale.

Run: python generate_sample_data.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
os.makedirs(OUTPUT_DIR, exist_ok=True)


def _cash_series(start, revenues, cogs_pct, opex_list):
    """Compute quarter-end cash balances from a starting balance and quarterly P&L flows."""
    cash, balance = [], start
    for rev, opex in zip(revenues, opex_list):
        balance += rev - (rev * cogs_pct) - opex
        cash.append(round(balance))
    return cash


# ─────────────────────────────────────────────────────────────────────────────
# Company 1 · CloudHive  (B2B SaaS — project-management platform)
# ─────────────────────────────────────────────────────────────────────────────
# Label style : "FY23 Q1"  (clean, standard)
# Format      : No junk rows, standard column names, sheet "Monthly_PnL"
# Trend       : Steady 43% CAGR; turns cash-flow positive in FY25 Q3
# Status      : Green  (≈25 quarters runway)
# ─────────────────────────────────────────────────────────────────────────────
def make_cloudhive():
    quarters = [
        "FY23 Q1","FY23 Q2","FY23 Q3","FY23 Q4",
        "FY24 Q1","FY24 Q2","FY24 Q3","FY24 Q4",
        "FY25 Q1","FY25 Q2","FY25 Q3","FY25 Q4",
    ]
    revenues = [
        25_000_000, 28_000_000, 31_000_000, 35_000_000,
        38_000_000, 42_000_000, 47_000_000, 52_000_000,
        57_000_000, 62_000_000, 67_000_000, 73_000_000,
    ]
    cogs_pct = 0.22
    opex = [
        47_000_000, 47_500_000, 48_000_000, 48_500_000,
        49_000_000, 49_500_000, 50_000_000, 50_500_000,
        51_000_000, 51_500_000, 52_000_000, 52_500_000,
    ]
    cogs = [round(r * cogs_pct) for r in revenues]
    cash = _cash_series(500_000_000, revenues, cogs_pct, opex)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly_PnL"
    ws.append(["Date", "Revenue", "COGS", "Operating_Expenses", "Cash_Balance"])
    for h in ws[1]:
        h.font = Font(bold=True)
    for q, r, c, o, cb in zip(quarters, revenues, cogs, opex, cash):
        ws.append([q, r, c, o, cb])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    wb.save(os.path.join(OUTPUT_DIR, "company_1_cloudhive.xlsx"))
    print("OK  company_1_cloudhive.xlsx  (SaaS - Green runway)")


# ─────────────────────────────────────────────────────────────────────────────
# Company 2 · PayZen  (Fintech — embedded payments & BaaS)
# ─────────────────────────────────────────────────────────────────────────────
# Label style : "Q1-FY23"  (quarter before year, hyphen-separated)
# Format      : 2 junk rows at top; renamed cols; sheet "P&L"
# Trend       : Revenue plateauing at ~₹2.4 Cr/quarter; 10% CAGR
# Status      : Amber  (≈3.1 quarters runway)
# ─────────────────────────────────────────────────────────────────────────────
def make_payzen():
    quarters = [
        "Q1-FY23","Q2-FY23","Q3-FY23","Q4-FY23",
        "Q1-FY24","Q2-FY24","Q3-FY24","Q4-FY24",
        "Q1-FY25","Q2-FY25","Q3-FY25","Q4-FY25",
    ]
    revenues = [
        18_000_000, 19_000_000, 20_000_000, 21_000_000,
        21_500_000, 22_000_000, 22_500_000, 23_000_000,
        23_500_000, 24_000_000, 24_000_000, 24_000_000,
    ]
    cogs_pct = 0.48
    opex = [
        28_000_000, 28_500_000, 29_000_000, 29_000_000,
        29_000_000, 29_500_000, 29_500_000, 30_000_000,
        30_000_000, 30_500_000, 30_500_000, 31_000_000,
    ]
    cogs = [round(r * cogs_pct) for r in revenues]
    cash = _cash_series(275_000_000, revenues, cogs_pct, opex)

    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws.append(["PayZen Technologies Pvt Ltd — Quarterly Financial Report"])
    ws.append(["Prepared by: Accounts Team", "", "FY 2023-25 (Quarterly)"])
    ws.append(["Period", "Gross Revenue", "Cost of Revenue", "Total Opex", "Cash"])
    for h in ws[3]:
        h.font = Font(bold=True)
    ws["A1"].font = Font(bold=True, size=12)
    for q, r, c, o, cb in zip(quarters, revenues, cogs, opex, cash):
        ws.append([q, r, c, o, cb])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    wb.save(os.path.join(OUTPUT_DIR, "company_2_payzen.xlsx"))
    print("OK  company_2_payzen.xlsx     (Fintech - Amber runway)")


# ─────────────────────────────────────────────────────────────────────────────
# Company 3 · FreshBowl  (Consumer — quick-commerce / dark kitchens)
# ─────────────────────────────────────────────────────────────────────────────
# Label style : "FY2023 Q1"  (full 4-digit year)
# Format      : 3 junk rows; Cash column FIRST (before Revenue); sheet "Income Statement"
# Trend       : Revenue crawling up but 70% COGS + heavy opex bleeds cash fast
# Status      : RED  (≈1.7 quarters runway — critical)
# ─────────────────────────────────────────────────────────────────────────────
def make_freshbowl():
    quarters = [
        "FY2023 Q1","FY2023 Q2","FY2023 Q3","FY2023 Q4",
        "FY2024 Q1","FY2024 Q2","FY2024 Q3","FY2024 Q4",
        "FY2025 Q1","FY2025 Q2","FY2025 Q3","FY2025 Q4",
    ]
    revenues = [
         90_000_000,  95_000_000, 100_000_000, 105_000_000,
        108_000_000, 110_000_000, 112_000_000, 115_000_000,
        116_000_000, 118_000_000, 119_000_000, 120_000_000,
    ]
    cogs_pct = 0.70
    opex = [
        45_000_000, 45_500_000, 46_000_000, 46_500_000,
        46_500_000, 47_000_000, 47_000_000, 47_500_000,
        47_500_000, 48_000_000, 48_000_000, 48_500_000,
    ]
    cogs = [round(r * cogs_pct) for r in revenues]
    cash = _cash_series(195_000_000, revenues, cogs_pct, opex)

    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    ws.append(["FRESHBOWL INDIA PVT LTD"])
    ws.append(["Internal Management Accounts — CONFIDENTIAL"])
    ws.append(["Reporting period: FY 2023 to FY 2025 (Quarterly)"])
    ws.append(["Month", "Cash on Hand", "Total Revenue", "COGS", "Operating Costs"])
    for h in ws[4]:
        h.font = Font(bold=True)
    ws["A1"].font = Font(bold=True, size=13)
    for q, r, c, o, cb in zip(quarters, revenues, cogs, opex, cash):
        ws.append([q, cb, r, c, o])    # cash column intentionally first
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    wb.save(os.path.join(OUTPUT_DIR, "company_3_freshbowl.xlsx"))
    print("OK  company_3_freshbowl.xlsx  (Consumer - RED runway - critical)")


# ─────────────────────────────────────────────────────────────────────────────
# Company 4 · DataPilot  (SaaS — enterprise data observability)
# ─────────────────────────────────────────────────────────────────────────────
# Label style : "2023 Q1"  (no FY prefix, bare 4-digit year)
# Format      : Snake_case headers, no junk rows, sheet "Financials"
# Trend       : Strongest compounder (43% CAGR); profitable from FY25 Q1
# Status      : Green  (cash-flow positive — runway effectively infinite)
# ─────────────────────────────────────────────────────────────────────────────
def make_datapilot():
    quarters = [
        "2023 Q1","2023 Q2","2023 Q3","2023 Q4",
        "2024 Q1","2024 Q2","2024 Q3","2024 Q4",
        "2025 Q1","2025 Q2","2025 Q3","2025 Q4",
    ]
    revenues = [
         60_000_000,  67_000_000,  74_000_000,  82_000_000,
         91_000_000, 100_000_000, 110_000_000, 121_000_000,
        133_000_000, 146_000_000, 160_000_000, 175_000_000,
    ]
    cogs_pct = 0.19
    opex = [
        85_000_000, 86_000_000, 87_000_000, 88_000_000,
        89_000_000, 90_000_000, 91_000_000, 92_000_000,
        93_000_000, 95_000_000, 97_000_000, 99_000_000,
    ]
    cogs = [round(r * cogs_pct) for r in revenues]
    cash = _cash_series(700_000_000, revenues, cogs_pct, opex)

    wb = Workbook()
    ws = wb.active
    ws.title = "Financials"
    ws.append(["month_year", "net_revenue", "cost_of_goods", "total_opex", "ending_cash"])
    for q, r, c, o, cb in zip(quarters, revenues, cogs, opex, cash):
        ws.append([q, r, c, o, cb])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    wb.save(os.path.join(OUTPUT_DIR, "company_4_datapilot.xlsx"))
    print("OK  company_4_datapilot.xlsx  (SaaS - Green runway - profitable)")


# ─────────────────────────────────────────────────────────────────────────────
# Company 5 · StyleLoop  (Consumer — D2C fashion, India)
# ─────────────────────────────────────────────────────────────────────────────
# Label style : "Q1 FY23"  (quarter then year, space-separated)
# Format      : Data on 2nd sheet ("Monthly Data"), 1 junk row then headers
# Trend       : 24% CAGR but seasonal (Q4 festive spike masks a slow bleed)
# Status      : Amber  (≈2.8 quarters runway)
# ─────────────────────────────────────────────────────────────────────────────
def make_styleloop():
    quarters = [
        "Q1 FY23","Q2 FY23","Q3 FY23","Q4 FY23",
        "Q1 FY24","Q2 FY24","Q3 FY24","Q4 FY24",
        "Q1 FY25","Q2 FY25","Q3 FY25","Q4 FY25",
    ]
    revenues = [
         55_000_000,  50_000_000,  58_000_000,  85_000_000,   # FY23: festive Q4 spike
         60_000_000,  55_000_000,  65_000_000,  95_000_000,   # FY24
         68_000_000,  62_000_000,  72_000_000, 105_000_000,   # FY25
    ]
    cogs_pct = 0.58
    opex = [
        50_000_000, 50_000_000, 51_000_000, 53_000_000,
        51_000_000, 51_000_000, 52_000_000, 54_000_000,
        52_000_000, 52_000_000, 53_000_000, 55_000_000,
    ]
    cogs = [round(r * cogs_pct) for r in revenues]
    cash = _cash_series(340_000_000, revenues, cogs_pct, opex)

    wb = Workbook()
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov["A1"] = "StyleLoop India"
    ws_ov["A2"] = "D2C Fashion Platform — Management Accounts FY23-FY25"
    ws_ov["A3"] = "Contact: cfo@styleloop.in"
    ws_ov["A5"] = "See 'Monthly Data' sheet for P&L figures."
    ws_ov["A1"].font = Font(bold=True, size=14)
    ws_ov["A5"].font = Font(italic=True)

    ws_data = wb.create_sheet("Monthly Data")
    ws_data.append(["StyleLoop — Quarterly P&L Summary (FY23-FY25)"])   # junk row
    ws_data.append(["Reporting_Month", "Sales_Revenue", "Product_COGS",
                    "SG&A_Expenses", "Bank_Balance"])
    for h in ws_data[2]:
        h.font = Font(bold=True)
    for q, r, c, o, cb in zip(quarters, revenues, cogs, opex, cash):
        ws_data.append([q, r, c, o, cb])
    for col in ws_data.columns:
        ws_data.column_dimensions[col[0].column_letter].width = 22
    wb.save(os.path.join(OUTPUT_DIR, "company_5_styleloop.xlsx"))
    print("OK  company_5_styleloop.xlsx  (Consumer - Amber runway)")


if __name__ == "__main__":
    print(f"Writing quarterly sample data to: {OUTPUT_DIR}\n")
    make_cloudhive()
    make_payzen()
    make_freshbowl()
    make_datapilot()
    make_styleloop()
    print("\nDone - 5 quarterly files created.")
