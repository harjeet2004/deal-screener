"""
mis_report.py  —  Portfolio MIS Dashboard Generator  (quarterly edition)

Pipeline (per .xlsx file in sample_data/):
  1. Detect the sheet with the most non-empty rows (handles data on sheet 2).
  2. Score rows on financial keywords to find the header row (skips junk rows).
  3. Map column names to the standard schema: quarter | revenue | cogs | opex | cash
     COGS is claimed before revenue so "Cost of Revenue" isn't grabbed as revenue.
  4. Parse quarter labels in any of 5 observed styles:
       "FY23 Q1"   "Q1-FY23"   "FY2023 Q1"   "2023 Q1"   "Q1 FY23"
  5. Compute per-company metrics (all in quarters, not months):
       QoQ revenue growth  — Q_n vs Q_{n-1}
       YoY revenue growth  — Q_n vs Q_{n-4}  (valid from Q5 onward; N/A earlier)
       Gross margin %      — (revenue − COGS) / revenue, latest quarter
       CAGR                — (Q12_rev / Q1_rev)^(1/3) − 1  over 3 years
       Quarterly burn      — mean(COGS + opex − revenue) across all quarters; 0 if profitable
       Runway (quarters)   — current cash / quarterly burn
       Status              — Critical < 2 qtrs | Watch 2–4 qtrs | Healthy > 4 qtrs
  6. Format all money in Indian Rupee style: ₹1,23,45,678
  7. Write MIS_Dashboard.xlsx — one summary row per company,
     bold navy header, frozen top row, descriptive RAG status cell.

Run: python3 mis_report.py
"""

import os
import re
import glob

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "sample_data")
OUTPUT   = os.path.join(BASE_DIR, "MIS_Dashboard.xlsx")

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
STRIPE_FILL = PatternFill("solid", fgColor="EEF2F7")
RED_FILL    = PatternFill("solid", fgColor="FF4C4C")
AMBER_FILL  = PatternFill("solid", fgColor="FFC000")
GREEN_FILL  = PatternFill("solid", fgColor="70AD47")
STATUS_FILL = {
    "Critical  (<2 qtrs)": RED_FILL,
    "Watch  (2-4 qtrs)":   AMBER_FILL,
    "Healthy  (>4 qtrs)":  GREEN_FILL,
}


# ─────────────────────────────────────────────────────────────────────────────
# Indian Rupee formatter
# ─────────────────────────────────────────────────────────────────────────────

def fmt_inr(amount):
    """
    Format a number in the Indian numbering system with ₹ symbol.
    Groups the last 3 digits, then groups of 2 from the right.
      2345600  →  ₹23,45,600
      12500000 →  ₹1,25,00,000
    Raw numbers are never modified; this is for display only.
    """
    if not isinstance(amount, (int, float)):
        return str(amount)
    neg    = amount < 0
    amount = abs(int(round(amount)))
    s      = str(amount)
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]
    return ("-" if neg else "") + "₹" + result   # ₹ = U+20B9


# ─────────────────────────────────────────────────────────────────────────────
# 1.  Sheet detection
# ─────────────────────────────────────────────────────────────────────────────

def find_data_sheet(wb):
    """Pick the worksheet with the most non-empty rows."""
    best, best_count = wb.active, 0
    for name in wb.sheetnames:
        ws    = wb[name]
        count = sum(
            1 for row in ws.iter_rows(values_only=True)
            if any(c is not None for c in row)
        )
        if count > best_count:
            best_count, best = count, ws
    return best


# ─────────────────────────────────────────────────────────────────────────────
# 2.  Header-row detection
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_KEYWORDS = {
    "revenue", "sales", "cogs", "cost", "opex", "operating",
    "cash", "balance", "bank", "period", "month", "quarter",
    "expenses", "sg&a", "gross", "date",
}

def find_header_row_index(ws):
    """Return 0-based index of the first row with 3+ financial keywords."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        text = " ".join(str(c).lower() for c in row if c is not None)
        if sum(1 for kw in _HEADER_KEYWORDS if kw in text) >= 3:
            return idx
    return 0


# ─────────────────────────────────────────────────────────────────────────────
# 3.  Column mapping  (unchanged from monthly version — logic is format-agnostic)
# ─────────────────────────────────────────────────────────────────────────────

def _first_col(headers_lower, includes, excludes=None):
    excludes = excludes or []
    for i, h in enumerate(headers_lower):
        if any(inc in h for inc in includes) and not any(exc in h for exc in excludes):
            return i
    return None


def map_columns(raw_headers):
    h = [str(x).lower().strip().replace("_", " ") if x else "" for x in raw_headers]

    cogs = _first_col(h, includes=["cogs", "cost of goods", "cost of revenue", "product cogs"])
    rev  = _first_col(h, includes=["revenue", "sales"], excludes=["cost", "cogs"])
    opex = _first_col(h, includes=["opex", "sg&a", "operating expense", "operating cost",
                                    "total opex", "sga"])
    if opex is None:
        opex = _first_col(h, includes=["expense"], excludes=["revenue", "cogs", "cost of"])
    cash = _first_col(h, includes=["cash", "bank", "balance"])
    date = _first_col(h, includes=["date", "month", "quarter", "period", "year"])

    return {"date": date, "revenue": rev, "cogs": cogs, "opex": opex, "cash": cash}


# ─────────────────────────────────────────────────────────────────────────────
# 4.  Quarter-label parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_quarter(val):
    """
    Parse any of the 5 observed quarter-label styles into a sortable (year, qtr) tuple.
    Handles: "FY23 Q1", "Q1-FY23", "FY2023 Q1", "2023 Q1", "Q1 FY23".
    Returns None for blank or unrecognisable values.
    """
    if val is None:
        return None
    s = str(val).strip().upper()
    year_m = re.search(r"(?:FY)?(\d{2,4})", s)
    qtr_m  = re.search(r"Q(\d)", s)
    if not year_m or not qtr_m:
        return None
    year = int(year_m.group(1))
    if year < 100:
        year += 2000
    return (year, int(qtr_m.group(1)))


def quarter_label(tup):
    """(2023, 1) → 'FY23 Q1' for display."""
    return f"FY{str(tup[0])[2:]} Q{tup[1]}"


# ─────────────────────────────────────────────────────────────────────────────
# 5.  Load and normalise one company file
# ─────────────────────────────────────────────────────────────────────────────

def load_company(filepath, fname=None):
    """
    Parse one Excel file (path string or BytesIO) and return a clean DataFrame:
    [quarter_key, quarter_label, revenue, cogs, opex, cash]
    sorted chronologically. Returns None on parse failure.
    """
    if fname is None:
        fname = os.path.basename(filepath) if isinstance(filepath, str) else "uploaded_file"
    wb    = openpyxl.load_workbook(filepath, data_only=True)
    ws    = find_data_sheet(wb)

    all_rows  = list(ws.iter_rows(values_only=True))
    hdr_idx   = find_header_row_index(ws)
    col_map   = map_columns(all_rows[hdr_idx])

    missing = [k for k, v in col_map.items() if v is None]
    if missing:
        print(f"    [WARN] Cannot map {missing} in {fname} — skipping")
        return None

    records = []
    for row in all_rows[hdr_idx + 1:]:
        qkey = parse_quarter(row[col_map["date"]])
        if qkey is None:
            continue
        try:
            records.append({
                "quarter_key":   qkey,
                "quarter_label": quarter_label(qkey),
                "revenue": float(row[col_map["revenue"]] or 0),
                "cogs":    float(row[col_map["cogs"]]    or 0),
                "opex":    float(row[col_map["opex"]]    or 0),
                "cash":    float(row[col_map["cash"]]    or 0),
            })
        except (TypeError, ValueError):
            continue

    if not records:
        print(f"    [WARN] No data rows in {fname} — skipping")
        return None

    df = (pd.DataFrame(records)
            .sort_values("quarter_key")
            .reset_index(drop=True))
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 6.  Metric computation
# ─────────────────────────────────────────────────────────────────────────────

def _pct(value, sign=False):
    if value is None:
        return "N/A"
    return (f"+{value*100:.1f}%" if (sign and value >= 0) else f"{value*100:.1f}%")


def compute_metrics(df, company_name):
    """
    Compute all MIS metrics from a normalised quarterly DataFrame.
    Returns a flat dict containing:
      - display-ready string values for the Excel dashboard
      - raw numeric values (prefixed _) for the Streamlit app and charts
    """
    last = df.iloc[-1]
    prev = df.iloc[-2] if len(df) >= 2 else None

    # ── QoQ revenue growth: Q_n vs Q_{n-1} ──────────────────────────────────
    qoq = None
    if prev is not None and prev["revenue"] != 0:
        qoq = (last["revenue"] - prev["revenue"]) / prev["revenue"]

    # ── YoY revenue growth: Q_n vs same quarter prior year (Q_{n-4}) ─────────
    yoy_str = "N/A (< 4 qtrs)"
    yoy_raw = None
    if len(df) >= 5:
        same_qtr_py = df.iloc[-5]
        if same_qtr_py["revenue"] != 0:
            yoy_raw = (last["revenue"] - same_qtr_py["revenue"]) / same_qtr_py["revenue"]
            yoy_str = _pct(yoy_raw, sign=True)

    # ── Gross margin %: latest quarter ───────────────────────────────────────
    gm_raw = None
    if last["revenue"] != 0:
        gm_raw = (last["revenue"] - last["cogs"]) / last["revenue"]

    # ── CAGR across 3 years ──────────────────────────────────────────────────
    # CAGR is the right long-horizon measure for a family office because it
    # smooths out quarter-to-quarter noise and expresses growth as a single
    # annualised rate — directly comparable across companies and asset classes.
    cagr_raw = None
    if len(df) >= 2 and df.iloc[0]["revenue"] > 0:
        n_years  = (df.iloc[-1]["quarter_key"][0] - df.iloc[0]["quarter_key"][0]
                    + (df.iloc[-1]["quarter_key"][1] - df.iloc[0]["quarter_key"][1]) / 4)
        n_years  = max(n_years, 0.25)   # guard against division by zero
        cagr_raw = (last["revenue"] / df.iloc[0]["revenue"]) ** (1 / n_years) - 1

    # ── Quarterly burn: average net cash outflow per quarter ─────────────────
    df = df.copy()
    df["net_outflow"] = df["cogs"] + df["opex"] - df["revenue"]
    avg_outflow    = df["net_outflow"].mean()
    quarterly_burn = max(avg_outflow, 0.0)   # 0 if cash-flow positive on average

    # ── Runway in quarters ───────────────────────────────────────────────────
    current_cash  = last["cash"]
    runway_qtrs   = (current_cash / quarterly_burn) if quarterly_burn > 0 else float("inf")

    # ── Status (thresholds in quarters) ──────────────────────────────────────
    if runway_qtrs > 4:
        status = "Healthy  (>4 qtrs)"
    elif runway_qtrs >= 2:
        status = "Watch  (2-4 qtrs)"
    else:
        status = "Critical  (<2 qtrs)"

    return {
        # ── Display values for Excel ──────────────────────────────────────
        "Company":            company_name,
        "Latest Quarter":     last["quarter_label"],
        "Revenue":            fmt_inr(last["revenue"]),
        "QoQ Rev Growth":     _pct(qoq, sign=True),
        "YoY Rev Growth":     yoy_str,
        "Gross Margin":       _pct(gm_raw),
        "Revenue CAGR (3Y)":  _pct(cagr_raw),
        "Qtrly Burn":         fmt_inr(quarterly_burn) if quarterly_burn > 0 else "Profitable",
        "Cash":               fmt_inr(current_cash),
        "Runway (qtrs)":      f"{runway_qtrs:.1f}" if quarterly_burn > 0 else "inf",
        "Status":             status,
        # ── Raw values for Streamlit app and charts ───────────────────────
        "_revenues":          df["revenue"].tolist(),
        "_cogs":              df["cogs"].tolist(),
        "_opex":              df["opex"].tolist(),
        "_cash":              df["cash"].tolist(),
        "_quarters":          df["quarter_label"].tolist(),
        "_qoq_raw":           qoq,
        "_yoy_raw":           yoy_raw,
        "_gm_raw":            gm_raw,
        "_cagr_raw":          cagr_raw,
        "_burn_raw":          quarterly_burn,
        "_runway_raw":        runway_qtrs,
        "_latest_rev_raw":    int(last["revenue"]),
        "_latest_cash_raw":   int(current_cash),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 7.  Write MIS_Dashboard.xlsx
# ─────────────────────────────────────────────────────────────────────────────

_DISPLAY_COLS = [
    "Company", "Latest Quarter", "Revenue", "QoQ Rev Growth",
    "YoY Rev Growth", "Gross Margin", "Revenue CAGR (3Y)",
    "Qtrly Burn", "Cash", "Runway (qtrs)", "Status",
]
_COL_WIDTHS = [22, 15, 18, 16, 18, 14, 18, 18, 18, 14, 22]


def write_dashboard(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "MIS Dashboard"

    ws.append(_DISPLAY_COLS)
    for cell in ws[1]:
        cell.font      = Font(color="FFFFFF", bold=True, size=11)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for i, row_data in enumerate(rows, start=2):
        ws.append([row_data.get(c, "") for c in _DISPLAY_COLS])
        status    = row_data.get("Status", "")
        is_stripe = (i % 2 == 1)
        for col_idx, cell in enumerate(ws[i], start=1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_name = _DISPLAY_COLS[col_idx - 1]
            if col_name == "Status":
                if status in STATUS_FILL:
                    cell.fill = STATUS_FILL[status]
                    cell.font = Font(bold=True)
            elif is_stripe:
                cell.fill = STRIPE_FILL

    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(OUTPUT)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def company_name_from_path(path):
    stem  = os.path.splitext(os.path.basename(path))[0]
    parts = [p for p in stem.split("_") if not p.isdigit() and p.lower() != "company"]
    return " ".join(p.capitalize() for p in parts)


def main():
    files = sorted(
        f for f in glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    )
    if not files:
        print(f"No .xlsx files found in {DATA_DIR}")
        return

    print(f"Processing {len(files)} quarterly files ...\n")
    all_metrics = []
    for filepath in files:
        name = company_name_from_path(filepath)
        print(f"  {name}  [{os.path.basename(filepath)}]")
        df = load_company(filepath)
        if df is None:
            continue
        m = compute_metrics(df, name)
        all_metrics.append(m)
        print(f"    Status: {m['Status']:<22}  Runway: {m['Runway (qtrs)']} qtrs  "
              f"QoQ: {m['QoQ Rev Growth']}  CAGR: {m['Revenue CAGR (3Y)']}")

    if not all_metrics:
        print("No data processed.")
        return

    print(f"\nWriting {OUTPUT} ...")
    write_dashboard(all_metrics)
    print("Done.")


if __name__ == "__main__":
    main()
